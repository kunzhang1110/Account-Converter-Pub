import os
import re
import pandas as pd
import openpyxl as pxl
from datetime import date, datetime
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side)
from account_rules import (CATEGORY_RULES,
                           MERCHANT_CATEGORY_RULES, MERCHANT_RULES)
from utils import (CONVERTED_PATH)

IS_DAYFIRST = True  # If use Australia date format dd/MM/yyyy
INPUT_FILE_PATH = './Transactions.csv'  # Nab Transactions file
NAB_SHEET_NAME = 'Nab Account'
NAB_IGNORE_CATEGORIES = ['Transfers out', 'Internal transfers', 'Transfers in']


def convert_nab_account(input_path: str = INPUT_FILE_PATH):
    ACCOUNT_COLS = ['Date', 'Details', 'AUD', 'CNY', 'USD', 'Category',
                    'Notes', 'Info']

    def get_date(row):
        try:
            dateObj = datetime.strptime(row['Date'], '%d %b %y').date()
        except:
            print(
                "Check Time:  {} does not match format \'%d-%b-%y\''".format(row['Transaction Details']))
            return date(1900, 1, 1)
        # get date from Transaction Details
        match = re.search('(\d{2})\/(\d{2})', row['Transaction Details'])
        if match:
            month = int(match.group(2))
            day = int(match.group(1))
            if month == 12 and dateObj.month == 1:  # crossing two years
                return date(dateObj.year-1, month, day)
            return date(dateObj.year, month, day)
        else:
            return dateObj

    def get_details(merchant_name):
        if merchant_name in MERCHANT_RULES:
            return MERCHANT_RULES[merchant_name]

        if (merchant_name) == "":
            return ""
        else:
            match = re.search('([^\(\)]+)(?:\(.*)?', merchant_name)
            return match.group(1).strip()

    def get_category(row):
        if row['Merchant Name'] in MERCHANT_CATEGORY_RULES:
            return MERCHANT_CATEGORY_RULES[row['Merchant Name']]
        if row['Category'] in CATEGORY_RULES:
            return CATEGORY_RULES[row['Category']]
        if row['Category'] == 'Uncategorised':
            return ''
        return row['Category']

    transactions = pd.read_csv(input_path)
    transactions.fillna("", inplace=True)
    transactions = transactions[~transactions['Category'].isin(
        NAB_IGNORE_CATEGORIES)]

    global account_df
    account_df = pd.DataFrame(columns=ACCOUNT_COLS)
    account_df['Date'] = transactions.loc[:, [
        'Date', 'Transaction Details']].apply(get_date, axis=1)
    account_df[['AUD', 'Info']] = transactions.loc[:,
                                                   ['Amount', 'Transaction Details']]
    account_df['Details'] = transactions.loc[:,
                                             'Merchant Name'].apply(get_details)
    account_df['Category'] = transactions.loc[:, [
        'Merchant Name', 'Category']].apply(get_category, axis=1)

    def condition_sum(col):
        if col.dtypes == 'float64':
            return sum(col.values)
        else:
            if col.name == 'Info':
                return '; '.join(col.fillna('').values)

    account_df = account_df.groupby(['Date', 'Details', 'Category']).aggregate(
        condition_sum).reset_index().reindex(columns=ACCOUNT_COLS)

    account_df.sort_values(by=['Date'], ascending=True, inplace=True)


def save_to_output(filepath: str = CONVERTED_PATH):
    NAB_CONVERTED_COLUMNS_WIDTHS = [13, 20] + [10] * 3 + [13, 10, 20]
    POSITIVE_NUM_BGCOLOR = 'C6EFCE'
    NEGATIVE_NUM_BGCOLOR = 'FFC7CE'
    POSITIVE_NUM_FONTCOLOR = '006100'
    NEGATIVE_NUM_FONTCOLOR = '9C0006'
    BORDER_STYLE = Border(left=Side(style='thin', color='A6A6A6'),
                          right=Side(style='thin', color='A6A6A6'),
                          top=Side(style='thin', color='A6A6A6'),
                          bottom=Side(style='thin', color='A6A6A6'))
    HEADER_STYLE = NamedStyle(name="header_style",
                              font=Font(color="FFFFFF", bold=True),
                              fill=PatternFill("solid", fgColor="000000"),
                              alignment=Alignment(
                                  horizontal="center", vertical="center"),
                              border=BORDER_STYLE)

    def set_header_style(ws):
        for row in ws.iter_rows(max_row=1):
            for cell in row:
                cell.style = HEADER_STYLE

    def set_columns_widths(ws, column_widths):
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

    if not os.path.exists(filepath):
        wb = pxl.Workbook()
        ws = wb.active
        ws.title = NAB_SHEET_NAME
    else:
        wb = pxl.load_workbook(filepath)
        if NAB_SHEET_NAME in wb.sheetnames:
            wb.remove(wb[NAB_SHEET_NAME])
            # clear custom named styles
            wb._named_styles = [wb._named_styles[0]]
        ws = wb.create_sheet(NAB_SHEET_NAME)

    for r in dataframe_to_rows(account_df, index=False, header=True):
        ws.append(r)

    set_header_style(ws)
    set_columns_widths(ws, NAB_CONVERTED_COLUMNS_WIDTHS)

    for row in ws.iter_rows(min_row=2):
        for i in range(len(row)):  # entire row
            row[i].border = BORDER_STYLE
            if i in [2, 3, 4] and row[i].value is not None:
                if float(row[i].value) > 0:
                    row[i].fill = PatternFill(
                        "solid", fgColor=POSITIVE_NUM_BGCOLOR)
                    row[i].font = Font(color=POSITIVE_NUM_FONTCOLOR)
                elif float(row[i].value) < 0:
                    row[i].fill = PatternFill(
                        "solid", fgColor=NEGATIVE_NUM_BGCOLOR)
                    row[i].font = Font(color=NEGATIVE_NUM_FONTCOLOR)
        row[0].number_format = "dd/MM/yyyy" if IS_DAYFIRST else "yyyy-MM-dd"  # Date

    wb.save(filename=CONVERTED_PATH)
    print("Output saved to {}".format(CONVERTED_PATH))


convert_nab_account()
save_to_output()
print("Convert completed.")
